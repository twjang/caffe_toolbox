#!/usr/bin/env python
from   __future__ import print_function

from settings import CAFFE_ROOT

import sys
import os
sys.path.append(os.path.join(CAFFE_ROOT, 'python'))

import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import random
import base64
import time
import scipy as sp
import scipy.io
import cv2
import h5py
import statsmodels.api as sm

COMMON_CODE = """
from   __future__ import print_function
import sys
import os
sys.path.append('""" +os.path.join(CAFFE_ROOT, 'python')+ """')

import caffe
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cm as cm 
import base64
import cv2
import h5py
import statsmodels.api as sm
import random
import time
import sklearn
import sklearn.manifold
from   sklearn.manifold import TSNE


caffe.set_mode_gpu()

def rand_cmap(nlabels, type='bright', first_color_black=True, last_color_black=False, verbose=True):
    \"\"\"
    Creates a random colormap to be used together with matplotlib. Useful for segmentation tasks
    :param nlabels: Number of labels (size of colormap)
    :param type: 'bright' for strong colors, 'soft' for pastel colors
    :param first_color_black: Option to use first color as black, True or False
    :param last_color_black: Option to use last color as black, True or False
    :param verbose: Prints the number of labels and shows the colormap. True or False
    :return: colormap for matplotlib
    \"\"\"
    from matplotlib.colors import LinearSegmentedColormap
    import colorsys
    import numpy as np

    if type not in ('bright', 'soft'):
        print ('Please choose "bright" or "soft" for type')
        return

    if verbose:
        print('Number of labels: ' + str(nlabels))

    # Generate color map for bright colors, based on hsv
    if type == 'bright':
        randHSVcolors = [(np.random.uniform(low=0.0, high=1),
                          np.random.uniform(low=0.2, high=1),
                          np.random.uniform(low=0.9, high=1)) for i in xrange(nlabels)]

        # Convert HSV list to RGB
        randRGBcolors = []
        for HSVcolor in randHSVcolors:
            randRGBcolors.append(colorsys.hsv_to_rgb(HSVcolor[0], HSVcolor[1], HSVcolor[2]))

        if first_color_black:
            randRGBcolors[0] = [0, 0, 0]

        if last_color_black:
            randRGBcolors[-1] = [0, 0, 0]

        random_colormap = LinearSegmentedColormap.from_list('new_map', randRGBcolors, N=nlabels)

    # Generate soft pastel colors, by limiting the RGB spectrum
    if type == 'soft':
        low = 0.6
        high = 0.95
        randRGBcolors = [(np.random.uniform(low=low, high=high),
                          np.random.uniform(low=low, high=high),
                          np.random.uniform(low=low, high=high)) for i in xrange(nlabels)]

        if first_color_black:
            randRGBcolors[0] = [0, 0, 0]

        if last_color_black:
            randRGBcolors[-1] = [0, 0, 0]
        random_colormap = LinearSegmentedColormap.from_list('new_map', randRGBcolors, N=nlabels)

    # Display colorbar
    if verbose:
        from matplotlib import colors, colorbar
        from matplotlib import pyplot as plt
        fig, ax = plt.subplots(1, 1, figsize=(15, 0.5))

        bounds = np.linspace(0, nlabels, nlabels + 1)
        norm = colors.BoundaryNorm(bounds, nlabels)

        cb = colorbar.ColorbarBase(ax, cmap=random_colormap, norm=norm, spacing='proportional', ticks=None,
                                   boundaries=bounds, format='%1i', orientation=u'horizontal')

    return random_colormap

## Taken from http://nbviewer.ipython.org/github/bvlc/caffe/blob/master/examples/filter_visualization.ipynb

def tsne(data, label):
    model = TSNE(n_components=2, random_state = 0)
    print ("tsne preparing..")
    data = model.fit_transform(data)
    print ("tsne done")
    all_lbl = list(set(list(label)))
    old2new = dict(zip(all_lbl, range(len(all_lbl))))
    new_cmap = rand_cmap(len(all_lbl), verbose=False)
    label = np.array([old2new[label[i]] for i in xrange(len(label))])
    plt.scatter(data[:,0], data[:,1], cmap=new_cmap, c=label, alpha=0.5, vmin=0, vmax=label.max())
    plt.show(block=False)
    

def tsnei(data_layer, label_layer, nsample):
    data = []
    lbl = []
    num_per_batch = bb(data_layer).shape[0]
    
    totnum = 0
    while totnum < nsample:
        totnum += num_per_batch
        go()
        data.append(bb(data_layer))
        lbl.append(bb(label_layer))
        print ("%d/%d" % (totnum, nsample), end='\\r')
        import sys;sys.stdout.flush()
    print("")
    data = np.concatenate(data)
    lbl = np.concatenate(lbl)

    tsne(data, lbl)


def cdf(data):
    tmp = np.ma.masked_array(data,np.isnan(data))
    tmp = tmp.reshape((tmp.size,))
    x = np.linspace(tmp.min(), tmp.max())
    ecdf = sm.distributions.ECDF(tmp)
    y = ecdf(x)
    diff_y = - (np.hstack([[0],y]) - np.hstack([y,[0]]))[1:y.size] / (x[1]-x[0])
    diff_x = x[1:]
    plt.figure()
    plt.subplot(1,2,1)
    plt.step(x, y)
    plt.subplot(1,2,2)
    plt.step(diff_x, diff_y)
    plt.show(block=False)


def mkarr(data):
    data = np.array(data.data).copy().reshape(data.shape)
    return data

def show(data, padsize=1, padval=None, defcm=cm.Greys_r):
    first = True

    if isinstance(data, list):
        datalst = data
    else:
        if len(data.shape)==4:
            datalst = [data[i,:,:,:] for i in xrange(data.shape[0])]
        else: datalst = [data]

    if len(datalst) > 10:
        print ("You're now plotting %d images. Continue?(y/N)" % len(datalst))
        sure = raw_input()
        if sure.strip().lower() != 'y':
            return

    for data in datalst:
        tmp = np.ma.masked_array(data,np.isnan(data))
        curmin = tmp.min()
        curmax = tmp.max()
        curzero = (curmin  + curmax) / 2.0
        if padval is None: pval = curzero
        else: pval = padval

        print ("min = %f" % curmin)
        print ("max = %f" % curmax)
        print ("contains NaN = "+ str(data.max() != curmax))
        
        del tmp
            
        
        if len(data.shape)==3:
            # force the number of filters to be square
            n = int(np.ceil(np.sqrt(data.shape[0])))
            padding = ((0, n ** 2 - data.shape[0]), (0, padsize), (0, padsize)) + ((0, 0),) * (data.ndim - 3)
            data = np.pad(data, padding, mode='constant', constant_values=(pval, pval))
            
            # tile the filters into an image
            data = data.reshape((n, n) + data.shape[1:]).transpose((0, 2, 1, 3) + tuple(range(4, data.ndim + 1)))
            data = data.reshape((n * data.shape[1], n * data.shape[3]) + data.shape[4:])


        if not first: plt.figure()
        plt.imshow(data,interpolation="nearest", cmap=defcm, vmin = data.min(), vmax = data.max())

        first = False

    plt.show(block=False)

def lslayer(curnet=None):
    global net
    if curnet is None: curnet = net
    for idx in xrange(len(curnet.layers)):
        layer_name = curnet._layer_names[idx]
        layer_type = curnet.layers[idx].type
        layer_blob_num = len(curnet.layers[idx].blobs)
        print ("layer[%d]:%s, %s" % (idx, layer_name, layer_type))
        for i in xrange(layer_blob_num):
            print ("    blob[%d]: %s" % (i, str(curnet.layers[idx].blobs[i].data.shape)))

def lsblob(curnet=None):
    global net
    if curnet is None: curnet = net
    for key in net.blobs.keys():
        print ("%s : %s" % (key, str(net.blobs[key].data.shape)))

def bb(*kargs):
    global net
    
    kargs = list(kargs)
    diff = False

    if 'diff' in kargs:
        kargs.remove('diff')
        diff = True
    if len(kargs) < 1: raise "See usage by typing hlp()"
    if isinstance(kargs[0], caffe._caffe.Net):
        curnet = kargs[0]
        try: idx1 = kargs[1]
        except IndexError: idx1 = None
        try: idx2 = kargs[2]
        except IndexError: idx2 = None
    else:
        curnet = net
        try: idx1 = kargs[0]
        except IndexError: idx1 = None
        try: idx2 = kargs[1]
        except IndexError: idx2 = None

    if not idx2 is None:
        if diff: return net.layers[idx1].blobs[idx2].diff
        else: return net.layers[idx1].blobs[idx2].data
    else:
        if diff: return net.blobs[idx1].diff
        else: return net.blobs[idx1].data

def dd(*kargs):
    kargs=list(kargs)
    kargs.append('diff')
    return bb(*kargs)

def go(*kargs): 
    if len(kargs) == 0:
        use_net=None
        data=None
    elif len(kargs) == 1:
        if not isinstance(kargs, caffe.Net):
            data = kargs[0]
        else: use_net = kargs[0]
    elif len(kargs) == 2:
        use_net = kargs[0]
        data = kargs[1]
    else:
        print ("See usage by typing hlp()")
        return
    global net
    if use_net is None: use_net = net
    if not data is None:
        pass
    use_net.forward()
    use_net.backward()

def fg(use_net=None):
    global net
    if use_net is None: use_net = net
    use_net.forward()

def bg(use_net=None):
    global net
    if use_net is None: use_net = net
    use_net.backward()

def bloblst(use_net=None):
    global net
    if use_net is None: use_net = net
    return net.blobs.keys()

def layerlst(curnet=None):
    res = []
    global net
    if curnet is None: curnet = net
    for idx in xrange(len(curnet.layers)):
        layer_name = curnet._layer_names[idx]
        layer_type = curnet.layers[idx].type
        layer_blob_num = len(curnet.layers[idx].blobs)
        if layer_blob_num > 0:
            res.append(idx)
    return res

def showrange(data):
    first = True
    if isinstance(data, list):
        datalst = data
    else:
        datalst = [data]

    for data in datalst:
        tmp = np.ma.masked_array(data,np.isnan(data))
        curmin = tmp.min()
        curmax = tmp.max()

        print ("min = %15e  max=%15e  NaN=%s" % (curmin, curmax, str(data.max() != curmax)))

def dbgsgd2(curnet=None):
    global net
    if curnet is None: curnet = net

    for i in curnet.blobs.keys():
        print (" == blob[%s]" % (i))
        
        data = curnet.blobs[i].data
        tmp = np.ma.masked_array(data,np.isnan(data))
        curmin = tmp.min()
        curmax = tmp.max()
        curmean = tmp.mean()
        curstd = tmp.std()
        print ("      sz: min=%+10e  max=%+10e  mean=%+10e  std=%+10e  NaN=%s" % (curmin, curmax, curmean, curstd, str(data.max() != curmax) ))

        data = curnet.blobs[i].diff
        tmp = np.ma.masked_array(data,np.isnan(data))
        curmin = tmp.min()
        curmax = tmp.max()
        curmean = tmp.mean()
        curstd = tmp.std()
        print ("    grad: min=%+10e  max=%+10e  mean=%+10e  std=%+10e  NaN=%s" % (curmin, curmax, curmean, curstd, str(data.max() != curmax) ))
        print ("")


def dbgsgd(curnet=None):
    global net
    if curnet is None: curnet = net
    for idx in xrange(len(curnet.layers)):
        layer_name = curnet._layer_names[idx]
        layer_type = curnet.layers[idx].type
        layer_blob_num = len(curnet.layers[idx].blobs)
        if layer_blob_num == 0: continue
        print ("")
        print ("layer[%d]:%s (%s)" % (idx, layer_name, layer_type))

        for i in xrange(layer_blob_num):
            print (" == blob[%d]" % (i))
            
            data = curnet.layers[idx].blobs[i].data
            tmp = np.ma.masked_array(data,np.isnan(data))
            curmin = tmp.min()
            curmax = tmp.max()
            curmean = tmp.mean()
            curstd = tmp.std()
            print ("      sz: min=%+10e  max=%+10e  mean=%+10e  std=%+10e  NaN=%s" % (curmin, curmax, curmean, curstd, str(data.max() != curmax) ))

            data = curnet.layers[idx].blobs[i].diff
            tmp = np.ma.masked_array(data,np.isnan(data))
            curmin = tmp.min()
            curmax = tmp.max()
            curmean = tmp.mean()
            curstd = tmp.std()
            print ("    grad: min=%+10e  max=%+10e  mean=%+10e  std=%+10e  NaN=%s" % (curmin, curmax, curmean, curstd, str(data.max() != curmax) ))
    

def hlp():
    print ("* hlp(): show help")
    print ("")
    print ("* go(net): run 1 iteration")
    print ("* fg(net): forward")
    print ("* bg(net): backward")
    print ("")
    print ("* mkarr(blob): make blob into np.array")
    print ("* show(data, padsize=1, padval=0): draw 4D-array whose shape is (numch, patch_h, patch_w)")
    print ("")
    print ("* lslayer(net[default by net]): list all layers")
    print ("* lsblob(net[default by net]): list all blobs")
    print ("")
    print ("* bb(net[default by net], blob_name): returns blob")
    print ("* bb(net[default by net], layeridx, blob_name): returns blob associated to layer[layeridx]")
    print ("* dd(net[default by net], blob_name): returns diff")
    print ("* dd(net[default by net], layeridx, blob_name): returns diff associated to layer[layeridx]")
    print ("")
    print ("* bloblst(net[default by net]): returns the list of all blob keys")
    print ("* layerlst(net[default by net]): returns the list of indices of all layers with weights")
    print ("")
    print ("* showrange(data): prints range of given matrices")
    print ("* dbgsgd(net[default by net]): debug learning rate")
    print ("* dbgsgd2(net[default by net]): debug learning rate")
    print ("" )
    print ("* cdf(data): draw sample cdf function")
"""

def randstr(length):
    return ''.join([chr(random.randint(0, 10) + ord('a')) for i in xrange(length)])

def do_extract(args):
    import caffe

    def ext(fname):
        return fname.split('.')[-1].lower()

    if len(args)<3:
        print ("Usage: [prototxt] [caffemodel] [output file name]")
        exit(-1)

    fname_prototxt = args[0]
    fname_caffemodel = args[1]
    fname_output = args[2]

    net = caffe.Net(fname_prototxt, fname_caffemodel, caffe.TEST)    
    layers = net.params.keys()
    final = {}
    for layer in layers:
        if len(net.params[layer]) > 0: final[layer+'_weight'] = net.params[layer][0].data
        if len(net.params[layer]) > 1: final[layer+'_bias']   = net.params[layer][1].data
        if len(net.params[layer]) > 2:
            for i in xrange(2, len(net.params[layer])):
                final[layer+'_param_%02d' % i] = net.params[layer][i].data

    fname_ext = ext(fname_output)
    if fname_ext in ['mat']:
        scipy.io.savemat(fname_output, final)
    elif fname_ext in ['hdf', 'h5', 'hdf5']:
        f = h5py.File(fname_output, 'w')
        for key in final.keys():
            f[key]=final[key]
        f.close()
    elif fname_ext in ['cc', 'c', 'cpp']:
        print (final.keys())
        f = file(fname_output, 'w')
        keys = final.keys()
        keys.sort()
        f.write('// ' + ' '.join(keys) + '\n') 
        for key in keys:
            f.write('extern int XXX_dim;\nextern int XXX_shape[];\nextern float XXX[];\n\n'.replace('XXX', key))
        f.write('\n\n')

        for key in keys:
            f.write('int %s_dim=%d;\n' % (key, len(final[key].shape)))
            f.write('int %s_shape[]={%s};\n' % (key, ', '.join([str(x) for x in list(final[key].shape)])))
            f.write('float %s[]={' % key)
            idx = 0
            tmp = final[key].reshape((final[key].size, ))
            for idx in xrange(final[key].size):
                if idx > 0: f.write(',')
                if idx % 25 == 0: f.write('\n')
                f.write('%f' % tmp[idx])
            f.write('\n};\n\n')
        f.close()

def do_clean(args):
    os.system("rm -rf ./snapshot/*")
    os.system("rm -f ./train.log")

def do_dataset(args):
    if len(args) < 1:
        print ("Usage: [lmdb path to inspect] [img per row=10] [img per col=10]")
        return 

    import lmdb
    import caffe.proto.caffe_pb2
    from caffe.proto.caffe_pb2 import Datum

    rowsz = 10
    colsz = 10
    if len(args) > 1: rowsz = int(args[1])
    if len(args) > 2: colsz = int(args[2])

    env = lmdb.open(args[0], readonly=True)
    print ("%d entries" % env.stat()['entries'] )
    try:
        with env.begin() as txn:
            with txn.cursor() as curs:
                while True:
                    stack = []
                    for key, value in curs:
                        d = Datum.FromString(value)
                        c = d.channels
                        h = d.height
                        w = d.width
                        lbl = d.label
                        im = np.fromstring(d.data, dtype=np.uint8).reshape(c, h, w)
                        stack.append((key, lbl, im))
                        if len(stack) == rowsz * colsz:
                            break

                    totalim = np.zeros((c, (h+1)*rowsz, (w+1)*colsz))
                    idx = 0
                    for key, lbl, im in stack:
                        x = idx % colsz
                        y = (idx - x) / colsz
                        x *= (w+1)
                        y *= (h+1)
                        idx += 1
                        totalim[:, y:y+h, x:x+w] = im

                    totalim = totalim.transpose(1,2,0) / 255.0
                    cv2.imshow("img", totalim)
                    pressed = cv2.waitKey(0)
                    if pressed == -1 or pressed == 1048603: break
    except KeyboardInterrupt:
        env.close()
        return
    env.close()



def do_pltacc(args):
    plt.ion()
    fname = args[0]

    if len(args)>1: recent_entry = int(args[1])
    else: recent_entry = None

    if len(args)>2: smooth = int(args[2])
    else: smooth = None

    while True:
        f=file(fname, 'r')
        lines=f.readlines()
        f.close()

        iter_indices = []
        val = []
        idx = 0 
        for line in lines:
            if line.find('Test net') > -1 and line.find('accuracy') > -1:
                v = line.strip().split('accuracy = ')[1]
                iter_indices.append(float(idx))
                val.append(float(v))
                idx += 1
        y = np.array(val)
        x = np.array(iter_indices)
        if not recent_entry is None:
            y = y[-recent_entry:]
            x = x[-recent_entry:]
        plt.clf()
        plt.plot(x, y)
        if not smooth is None:
            yy = moving_average(y, smooth)
            xx = moving_average(x, smooth)
            plt.plot(xx, yy)
        plt.pause(2.0)

def moving_average(a, n=3) :
    ret = np.cumsum(np.array(a), dtype=float)
    ret[n:] = ret[n:] - ret[:-n]
    return ret[n - 1:] / n


def do_pltmulti(args):
    from matplotlib.font_manager import FontProperties
    fontP = FontProperties()
    fontP.set_size('x-small')

    def draw_acc(fname):
        f=file(fname, 'r')
        lines=f.readlines()
        f.close()
        ## draw accuracy
        iter_indices = []
        val = []
        idx = 0 
        for line in lines:
            if line.find('Test net') > -1 and line.find('accuracy') > -1:
                v = line.strip().split('accuracy = ')[1]
                iter_indices.append(float(idx))
                val.append(float(v))
                idx += 1
        y = np.array(val)
        x = np.array(iter_indices)

        return (x, y)

    def draw_loss(fname):
        f=file(fname, 'r')
        lines=f.readlines()
        f.close()
        ## draw loss        
        iter_indices = []
        val = []
        for line in lines:
            if line.find('Iteration') > -1 and line.find('loss') > -1:
                v = line.strip().split('loss = ')[1]
                i = line.strip().split('Iteration ')[1].split(',')[0]
                iter_indices.append(float(i))
                val.append(float(v))
        y = np.array(val)
        x = np.array(iter_indices)

        return (x, y)

    #plt.ion()
    fig = plt.gcf()
    plt.clf()

    plt.subplot(1,2,1) 
    line_acc=[]
    for fname in args:
        if os.path.isdir(fname):
            fname = os.path.join(fname, 'train.log')
        x, y=draw_acc(fname)
        p,=plt.plot(x, y, label=fname)
        line_acc.append(p)
    plt.legend(handles=line_acc, prop=fontP, loc='best')

    plt.subplot(1,2,2) 
    line_loss=[]
    for fname in args:
        if os.path.isdir(fname):
            fname = os.path.join(fname, 'train.log')
        x, y=draw_loss(fname)
        p,=plt.plot(x, y, label=fname)
        line_loss.append(p)
    plt.legend(handles=line_loss, prop=fontP, loc='best')
    plt.show()




def do_pltloss(args):
    plt.ion()
    if len(args) == 0: fname = '.'
    else: fname = args[0]

    if os.path.isdir(fname):
        fname = os.path.join(fname, 'train.log')

    if len(args)>1: recent_entry = int(args[1])
    else: recent_entry = None
    
    if recent_entry == -1: resent_entry = None

    if len(args)>2: smooth = int(args[2])
    else: smooth = None
    if smooth == -1: smooth = None
    
    if len(args)>3: title = args[3]
    else: title = "test acc. / trng loss"


    while True:
        f=file(fname, 'r')
        lines=f.readlines()
        f.close()

        ## draw accuracy
        iter_indices = []
        val = []
        idx = 0 
        for line in lines:
            if line.find('Test net') > -1 and line.find('accuracy') > -1:
                v = line.strip().split('accuracy = ')[1]
                iter_indices.append(float(idx))
                val.append(float(v))
                idx += 1
        y = np.array(val)
        x = np.array(iter_indices)
        if not recent_entry is None:
            y = y[-recent_entry:]
            x = x[-recent_entry:]
       
        fig = plt.gcf()
        fig.canvas.set_window_title(title)
        plt.clf()
        plt.subplot(1,2,1)
        line1 = plt.plot(x, y)

        if not smooth is None:
            yy = moving_average(y, smooth)
            xx = moving_average(x, smooth)
            line2 = plt.plot(xx, yy)
        
        
        if not smooth is None:
            plt.setp(line1, 'color', '#aaaaaa')
            plt.setp(line2, 'color', '#ff0000')


        ## draw loss        
        iter_indices = []
        val = []

        for line in lines:
            if line.find('Iteration') > -1 and line.find('loss') > -1:
                v = line.strip().split('loss = ')[1]
                i = line.strip().split('Iteration ')[1].split(',')[0]
                iter_indices.append(float(i))
                val.append(float(v))
        y = np.array(val)
        x = np.array(iter_indices)
        if not recent_entry is None:
            y = y[-recent_entry:]
            x = x[-recent_entry:]

        plt.subplot(1,2,2)
        line1 = plt.plot(x, y)

        if not smooth is None:
            yy = moving_average(y, smooth)
            xx = moving_average(x, smooth)
            line2=plt.plot(xx, yy)

        if not smooth is None:
            plt.setp(line1, 'color', '#aaaaaa')
            plt.setp(line2, 'color', '#ff0000')
        plt.pause(1.0)
        
def do_eval(args):
    import caffe
    caffe.set_mode_gpu()

    import openpyxl  
    if len(args)<2:
        print ("arg1=model.prototxt, arg2=model.caffemodel, arg3=output name, arg4=# of test examples(optional), arg5=report(optional), arg6=class_name_list.txt")
        return

    model_def = args[0]
    model_fname = args[1]
    output_layer = args[2].strip()
    if len(args) > 3: report_path = args[4]
    else: report_path = None


    net = caffe.Net(model_def, model_fname, caffe.TEST) 
    print ("loaded..")

    
    class_num = None 
    if net.blobs.has_key(output_layer):
        class_num = net.blobs[output_layer].data.shape[1]
    else:
        print ("layer %s is not found" % output_layer)
        print ("possible layers are: %s" % (', '.join(list(net._blob_names))))
        return 

    if len(args) >= 6:
        class_lbl = file(args[5], 'r').readlines()
        class_lbl = map(lambda x:x.strip().decode('utf-8'), class_lbl)
        class_lbl = filter(lambda x:len(x)>0, class_lbl)
    else:
        class_lbl = map(lambda x:u'(%d)' % x, range(class_num))

    def cls2char(cls):
        return class_lbl[int(cls)].encode('utf-8')

    def escape_cls2char(cls):
        res = ''
        for ch in class_lbl[int(cls)]:
            res += '&#%d;' % ord(ch)
        return res

    confusion_matrix = np.zeros([class_num,class_num])

    answers = []
    answers_by_cls = {}
    
    num_batch=net.blobs['label'].data.size
    num_iter=100 if len(args) < 4 else int(float(args[3])/num_batch + 0.999)

    for i in xrange(num_iter):
        net.forward()
        label = list(net.blobs['label'].data.astype(np.int64))
        result = list(net.blobs[output_layer].data.argmax(1))
        
        for lbl, out in zip(label, result):
            confusion_matrix[lbl, out] += 1.0

        for idx in xrange(len(result)):
            score_lst = list(net.blobs[output_layer].data[idx].copy())
            score_lst = zip(score_lst, range(len(score_lst)))
            score_lst.sort()
            score_lst.reverse()
            score_lst = score_lst[0:5]
            ent = (net.blobs['data'].data[idx].astype(np.float), int(label[idx]), int(result[idx]), score_lst)
            answers.append(ent)
        print ("%d/%d" % (i, num_iter), end='\r')
        import sys;sys.stdout.flush()
    print ("")

    for ent in answers:
        img, lbl, res, score_lst = ent
        if answers_by_cls.has_key(lbl):
            isok = 1 if lbl == res else 0
            answers_by_cls[lbl].append((isok, (img, res, score_lst)))
        else:
            isok = 1 if lbl == res else 0
            answers_by_cls[lbl] = [(isok, (img, res, score_lst)),]
    
    for key in answers_by_cls.keys():
        entries = answers_by_cls[key]
        entries.sort(key=lambda e: e[0])
        answers_by_cls[key] = entries

    if report_path:
        try: os.makedirs(report_path)
        except OSError: pass
        f = file(os.path.join(report_path, 'index.html'), 'w')
        f.write("""
            <!DOCTYPE html>
            <html>

            <frameset cols="25%,*">
            <frame src="menu.htm">
            <frame name="view">
            </frameset>

            </html>""")
        f.close()

        fmenu = file(os.path.join(report_path, 'menu.htm'), 'w')
        fmenu.write('<html><head></head><body>')
        
        # Let's do not care about the fucking xlsx file, as we have good mat file
        if False:
            book = openpyxl.Workbook(encoding="utf-8")
            sheet_rc = book.create_sheet(title='raw')
            sheet_p = book.create_sheet(title='percent')
            #remove default sheet
            for sheetname in book.get_sheet_names():
                if sheetname in ['raw', 'percent']: continue
                s=book.get_sheet_by_name(sheetname)
                book.remove_sheet(s)

            sheets=[sheet_rc, sheet_p]

            for idx in xrange(len(sheets)):
                sheets[idx].cell(row=0, column=0).value="Accuracy"
                sheets[idx].cell(row=0, column=1).value=np.trace(confusion_matrix) / confusion_matrix.sum() * 100.0
                sheets[idx].cell(row=1, column=0).value='GT \\ Output'

            for i in xrange(0, class_num):
                for idx in xrange(len(sheets)):
                    sheets[idx].cell(row=1,   column=i+1).value='%s(%d)' % ( cls2char(i), i)
                    sheets[idx].cell(row=i+2, column=0).value='%s(%d)' % ( cls2char(i), i)

            tot_cnt = confusion_matrix.sum()
            for gt_idx in xrange(0, class_num):
                for out_idx in xrange(0, class_num):
                    sheets[0].cell(row=gt_idx+2, column=out_idx+1).value=confusion_matrix[gt_idx, out_idx]
                    sheets[1].cell(row=gt_idx+2, column=out_idx+1).value=float(confusion_matrix[gt_idx, out_idx]) / float(tot_cnt)

            # freeze panes
            sheets[0].freeze_panes = sheets[0].cell('B3')
            sheets[1].freeze_panes = sheets[1].cell('B3')

            book.save(os.path.join(report_path, 'main.xlsx'))

        ## save confusion_matrix in a mat file 
        sp.io.savemat(os.path.join(report_path, 'confusion.mat'), {'confusion': np.array(confusion_matrix)})

        ## best class pairs to merge 
        sym = confusion_matrix - np.diag(np.diag(confusion_matrix))
        sym = sym + sym.T

        conf_lst = []
        n = sym.shape[0]
        for i in xrange(n):
            for j in xrange(i+1, n):
                conf_lst.append((sym[i, j], i, j))

        conf_lst.sort()
        conf_lst.reverse()
        
        fconf=file(os.path.join(report_path, 'confusion.txt'), 'wb')
        fconf.write('freq,freq_percent,cls_label1,cls1,cls_label2,cls2,accum_freq,unique_cls\n')
        accum_freq = 0
        clsset=set()
        for freq, i, j in conf_lst:
            accum_freq += freq
            clsset.add(i)
            clsset.add(j)
            line = u'%d,%f%%,%s,%d,%s,%d,%d,%d\n' % (freq, 100.0 * float(freq) / float(confusion_matrix.sum()), escape_cls2char(i), i, escape_cls2char(j),j, accum_freq, len(clsset))
            if freq == 0: break
            fconf.write(line)
        fconf.close()

        fmenu.write('<a href="./confusion.mat" target="view">Summary</a><br/>\n')
        try: os.makedirs(os.path.join(report_path, 'im_f'))
        except OSError: pass
        try: os.makedirs(os.path.join(report_path, 'im_t'))
        except OSError: pass
        for cls in answers_by_cls.keys():
            fmenu.write('<a href="./cls_%d.html" target="view">%d(%s)</a><br/>\n' % (cls, cls, escape_cls2char(cls)))
            idxcnt = 0
            idxfname = os.path.join(report_path, 'cls_%d.html' % cls)
            idxf = file(idxfname, 'w')
            idxf.write('<html><head>')
            idxf.write('<style>.item { border:1px solid #000; display:inline-block; width:150px;}\n')
            idxf.write('.green { color:#20EE20; font-weight:bold;}\n')
            idxf.write('.wrong { color:#EE2020; font-weight:bold;}\n')
            idxf.write('</style>')
            idxf.write('</head><body>')
            idxf.write('<h1>%d(%s)</h1>\n' % (cls, escape_cls2char(cls)))
            for isok, data in answers_by_cls[cls]:
                img, res, score_lst = data
                idxcnt += 1
                img -= img.min()
                img /= img.max()
                img = img[0]
                
                if isok: path_ok = 'im_t'
                else: path_ok = 'im_f'

                cv2.imwrite(os.path.join(report_path, path_ok,'%05d_%05d.png' % (cls, idxcnt)), img * 255)

                if res == cls: anscolor = 'green'
                else: anscolor = 'wrong'
                top5str = '<br/>\n'.join(['<span style="font-size:0.3em;color:#aaa;">%s-%f</span>' % (escape_cls2char(clsnum), score) for score, clsnum in score_lst])
                idxf.write('<div class="item"><img src="./%s/%05d_%05d.png" /><span class="%s">%s</span>(%s)<br/>%s</div>\n' % (path_ok, cls, idxcnt, anscolor, escape_cls2char(res), escape_cls2char(cls), top5str))
            idxf.write('</body></html>')
            idxf.close()

        fmenu.write('</body></html>') 
        fmenu.close()

        f = file(os.path.join(report_path, 'accuracy.txt'), 'w')
        f.write("accuracy: %f\n" % (np.trace(confusion_matrix) / confusion_matrix.sum()))
        f.close()

    print ("accuracy: %f" % (np.trace(confusion_matrix) / confusion_matrix.sum()))

def do_ipython(args):
    global CAFFE_ROOT
    code = """
import sys
import os
sys.path.append('""" + os.path.join(CAFFE_ROOT, 'python') + """')

import caffe
import numpy as np
import scipy as sp
import matplotlib.pyplot as plt
"""
    tmpfname = '/tmp/' +randstr(10) 
    f=file(tmpfname, 'w') 
    f.write(code)
    f.close()

    os.system('ipython -i %s' % tmpfname)
    os.system('rm -f %s' % tmpfname)

def do_mean(args):
    global CAFFE_ROOT
    os.system("""
        export CAFFE_PATH=\"""" + CAFFE_ROOT + """\"
        $CAFFE_PATH./build/tools/compute_image_mean -backend lmdb ./dataset/chinese_train_lmdb ./dataset/chinese_mean.binaryproto
        echo "Done."
        """)


def do_train(args):
    global CAFFE_ROOT
    if len(args) > 0: gpuid = args[0]
    else: gpuid = "all"
    os.system("""
        export CAFFE_PATH=\"""" + CAFFE_ROOT + """\"
        $CAFFE_PATH/./build/tools/caffe train -gpu %s --solver=solver.prototxt 2>&1 | tee -a train.log
        #./learn train ./charconf.ini  --solver=solver.prototxt 2>&1 | tee -a train.log
    """ % gpuid )



def do_resume(args):
    global CAFFE_ROOT
    if len(args)<1:
        print ("Please specify the .caffemodel or .solverstate file")
        return

    fname = args[0]
    if fname.endswith('.solverstate'):
        os.system("""
            export CAFFE_PATH=\"""" + CAFFE_ROOT + """\"
            $CAFFE_PATH/./build/tools/caffe train  --solver solver.prototxt --snapshot %s 2>&1 | tee -a train.log
        """% (fname))
    elif fname.endswith('.caffemodel'):
        os.system("""
            export CAFFE_PATH=\"""" + CAFFE_ROOT + """\"
            $CAFFE_PATH/./build/tools/caffe train  --solver solver.prototxt --weights %s 2>&1 | tee -a train.log
        """% (fname))

def do_bench(args):
    if len(args)<2:
        print ("Usage: [model prototxt or caffemodel] [output xlsx] [device:cpu if not specified. or gpuid or all(all gpu)]")
        return

    import shlex, subprocess    
    import openpyxl

    fname = args[0]
    outfname = args[1]
    if len(args) < 3: device = 'cpu'
    else: device = args[2].strip().lower()

    
    if device == 'cpu': gpuflag = ''
    else: gpuflag = '-gpu %s' % device
    
    if fname.endswith('.caffemodel'):
        cmd=("%s/./build/tools/caffe time -weights %s %s" % (CAFFE_ROOT, fname, gpuflag))
    elif fname.endswith('.prototxt'):
        cmd=("%s/./build/tools/caffe time -model %s %s" % (CAFFE_ROOT, fname, gpuflag))

    args = shlex.split(cmd)
    p = subprocess.Popen(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    print ("Benchmark started..")
    stdout, stderr = p.communicate()
    print (stderr)
    print ("Benchmark Done.")

    lst = stderr.split('\n')
    idx = 0
    for line in lst:
        if line.find('Average time per layer:') > -1: 
            break
        idx += 1

    lst = filter(lambda x: x.find('] ') > -1, lst[idx+1:])
    
    layers=[]
    forward_time={}
    backward_time={}

    total_forward = 0.0
    total_backward = 0.0

    lst = [ line.split('] ',1)[1].strip().split() for line in lst ]
    for line in lst:
        if not 'forward:' in line and not 'backward:' in line: continue
        layer_name = line[0]
        is_forward = (line[1].lower().find('forward') > -1)
        time = line[2]
        time_unit = line[3].replace('.', '')

        if not layer_name in layers:
            layers.append(layer_name)

        if is_forward:
            forward_time[layer_name] = time
            total_forward += float(time)
        else:
            backward_time[layer_name] = time
            total_backward += float(time)

    book = openpyxl.Workbook(encoding="utf-8")
    sheet = book.create_sheet(title='result')
    #remove default sheet
    for sheetname in book.get_sheet_names():
        if sheetname in ['result']: continue
        s=book.get_sheet_by_name(sheetname)
        book.remove_sheet(s)

    sheet.cell(row=0, column=1).value="forward(ms)"
    sheet.cell(row=0, column=2).value="forward(%)"
    sheet.cell(row=0, column=3).value="backward(ms)"
    sheet.cell(row=0, column=4).value="backward(%)"
    
    last_row = len(layers)+1
    row = 1
    for layer_name in layers:
        sheet.cell(row=row, column=0).value = layer_name
        if forward_time.has_key(layer_name):
            sheet.cell(row=row, column=1).value = forward_time[layer_name]
            sheet.cell(row=row, column=2).value = '=B%d/$B$%d' % (row+1, last_row+1)
        if backward_time.has_key(layer_name):
            sheet.cell(row=row, column=3).value = backward_time[layer_name]
            sheet.cell(row=row, column=4).value = '=D%d/$D$%d' % (row+1, last_row+1)
        row += 1

    sheet.cell(row=last_row, column=1).value = '=sum(B2:B%d)' % last_row
    sheet.cell(row=last_row, column=3).value = '=sum(D2:D%d)' % last_row

    # freeze panes
    sheet.freeze_panes = sheet.cell('B2')
    book.save(outfname)


def do_loadstate(args):
    global COMMON_CODE
    if len(args)<2:
        print ("usage: [ prototxt ] [ caffemodel ]  [optional:mode (default=test)]")
        print ("Please specify the .caffemodel file")
        return 

    tmpfname = '/tmp/' +randstr(10) 

    mode='TEST'
    if len(args) > 2:
        if args[2].lower().strip() in ['train', 'trng']:
            mode = 'TRAIN'
        elif args[2].lower().strip() in ['test', 'val']:
            mode = 'TEST'
        else:
            print ("Unrecognized mode: %s" % args[2])
            print ("valid candidates are train,trng,test,val")
            return

    code = COMMON_CODE + """
model_def = base64.b64decode('%s')
model_fname = base64.b64decode('%s')
self_fname = base64.b64decode('%s')
net = caffe.Net(model_def, model_fname, caffe.%s) 
del model_fname
del model_def

print (" *** net_param is set for you")
print (" *** to plot the weight, ")
print ("")
print ("d = mkarr(bb(2,0))")
print ("show(d)")

hlp()
try: os.remove(self_fname)
except OSError: pass

""" % (base64.b64encode(args[0]), base64.b64encode(args[1]), base64.b64encode(tmpfname), mode)
    f=file(tmpfname, 'w') 
    f.write(code)
    f.close()

    os.system('ipython -i %s' % tmpfname)
    os.system('rm -f %s' % tmpfname)




def do_pltblob(args):
    import caffe
    if len(args)>0: fname = args[1]
    else: fname = './dataset/chinese_mean.binaryproto'

    data = file(fname,'r').read()
    b=caffe.io.caffe_pb2.BlobProto()
    b.ParseFromString(data)
    arr = caffe.io.blobproto_to_array(b)

    c=arr.shape[1]
    h=arr.shape[2]
    w=arr.shape[3]

    res = np.hstack([arr[0, 0, :,:].reshape((h,w)), arr[0, 1, :,:].reshape((h,w)), arr[0, 2, :,:].reshape((h,w))])
    myarr = np.zeros((h, w, c))
    for idx in xrange(c): myarr[:,:,idx] = arr[0, idx, :, :].reshape((h, w))

    plt.subplot(211)
    plt.imshow(res)
    plt.subplot(212)
    plt.imshow(myarr)
    plt.show()

def do_draw(args):
    global CAFFE_ROOT
    if len(args)>0: fname = args[0]
    else: fname='./model_train.prototxt'

    tmpfname = '/tmp/'+randstr(10)+'.png'
    os.system(os.path.join(CAFFE_ROOT, 'python/draw_net.py') + ' %s %s' % (fname,tmpfname))
    os.system('shotwell %s' % tmpfname)
    try: os.remove(tmpfname)
    except OSError: pass

def do_pack(args):
    if (len(args) < 1):
        print ("usage: [pack dir name]")
        return
    
    d = args[0]
    try: os.makedirs(d)
    except OSError: pass
    
    os.system('cp ./snapshot/ "./%s/" -R' % d)
    os.system('cp ./model_train.prototxt "./%s/" ' % d)
    os.system('cp ./train.log "./%s/" ' % d)
    os.system('cp ./solver.prototxt "./%s/" ' % d)


def do_visevol(args):
    if (len(args) < 3):
        print ("usage: [model def] [snapshot dir] [output dir]")
        return
    
    model_def = args[0]
    model_dir = args[1]
    path = args[2]
    
    snapshots = []
    for c in os.listdir(model_dir):
        if c.endswith('.caffemodel'):
            d = c[:-len('.caffemodel')]
            iternum = int(d.split('_')[-1])
            snapshots.append((iternum, os.path.join(model_dir, c)))

    snapshots.sort()
    for iternum, model_fname in snapshots:
        steppath = os.path.join(path, 'iter_%d' % iternum)
        if not os.path.exists(steppath):
            do_visweight([model_def, model_fname, steppath, 'index.html' ])
    
    f = file(os.path.join(path, 'index.html'), 'w')
    f.write("""
        <!DOCTYPE html>
        <html>

        <frameset cols="25%,*">
        <frame src="menu.htm">
        <frame name="view">
        </frameset>

        </html>""")
    f.close()
    f = file(os.path.join(path, 'menu.htm'), 'w')
    f.write('<html><head></head><body>')
    for iternum, model_fname in snapshots:
        f.write('<a href="./iter_%d/index.html" target="view">%d</a><br/>\n' % (iternum, iternum))
    f.write('</body></html>') 
    f.close()

def do_evolmovie(args):
    if (len(args) < 3):
        print ("usage: [visevol output dir] [layer id] [output file]")
        return

    visevol_path = args[0]
    layerid = args[1]
    outfname = args[2]
    
    imglst = []
    for c in os.listdir(visevol_path):
        curdir = os.path.join(visevol_path, c)
        if c.startswith('iter_') and os.path.isdir(curdir):
            imgdir = os.path.join(curdir, 'imgs')
            imgpath = os.path.join(imgdir, layerid + '.jpg')
            iternum = int(c.replace('iter_', ''))
            if os.path.isfile(imgpath):
                imglst.append((iternum, imgpath))
            else:
                print ("No such layer id(%s)" % layerid)
                print ("Please use .. ")
                for f in os.listdir(imgdir):
                    if not f.endswith('.jpg'): continue
                    print (f.replace('.jpg', ''))
                return
    
    if len(imglst) == 0:
        print ("Please check your arguments: no images found")
        return

    imglst.sort()

    im = cv2.imread(imglst[0][1])
    isColor = 1
    fps = 15
    frameW = im.shape[1]
    frameH = im.shape[0]

    print ("%d x %d" % (frameH, frameW))

    fourcc = cv2.cv.CV_FOURCC('m', 'p', '4', 'v')
    writer = cv2.VideoWriter(outfname, fourcc, fps, (frameW, frameH)) 
    for iternum, imgf in imglst:
        im = cv2.imread(imgf)
        cv2.putText(im, '%10d' % iternum, (0, 12), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 0, 0)) 
        writer.write(im)

    cv2.destroyAllWindows()
    writer.release()
    writer = None
    print ("Done")


def do_visweight(args):
    import caffe
    if (len(args) < 3):
        print ("usage: [model def] [caffemodel] [output dir] [output fname]")
        return
    
    model_def = args[0]
    model_fname = args[1]
    path = args[2]
    imgpath = os.path.join(path, 'imgs')
    if len(args)>3: output_fname = os.path.join(path, args[3])
    else: output_fname = os.path.join(path, 'index.html')
    
    try: os.makedirs(path)
    except OSError: pass
    try: os.makedirs(imgpath)
    except OSError: pass

    ## Weight development

    net = caffe.Net(model_def, model_fname, caffe.TEST) 
    
    f = file(output_fname, 'w')
    f.write('<html>') 
    f.write('<head>') 
    f.write('</head>') 
    f.write('<body>') 



    for idx in xrange(len(net.layers)):
        layer_blob_num = len(net.layers[idx].blobs)
        layer_name = net._layer_names[idx]
        if layer_blob_num > 0:

            f.write('<h3>%s</h3>'% layer_name)
            weightdata = net.layers[idx].blobs[0].data

            if len(weightdata.shape)==4:
                imgfname = 'w%02d.jpg' % (idx)
                curweight_imgpath = os.path.join(imgpath, imgfname)
                curweight_relimgpath = os.path.join('imgs', imgfname)
                f.write('<img src="%s" />' % curweight_relimgpath)
                

                ### create weight image for each input channels
                gg = []
                
                for widx in xrange(weightdata.shape[1]):
                    data = weightdata [:,widx]

                    padsize = 1
                    padval = data.mean()

                    # force the number of filters to be square
                    n = int(np.ceil(np.sqrt(data.shape[0])))
                    padding = ((0, n ** 2 - data.shape[0]), (0, padsize), (0, padsize)) + ((0, 0),) * (data.ndim - 3)
                    data = np.pad(data, padding, mode='constant', constant_values=(0, padval))
                    
                    # tile the filters into an image
                    data = data.reshape((n, n) + data.shape[1:]).transpose((0, 2, 1, 3) + tuple(range(4, data.ndim + 1)))
                    data = data.reshape((n * data.shape[1], n * data.shape[3]) + data.shape[4:])

                    data -= data.min()
                    data /= data.max()

                    gg.append(data)


                gg = np.array(gg) 

                ### we create the final output image by remapping the collected weight images
                data = gg
                    
                padsize = 4
                padval = data.mean()

                # force the number of filters to be square
                n = int(np.ceil(np.sqrt(data.shape[0])))
                padding = ((0, n ** 2 - data.shape[0]), (0, padsize), (0, padsize)) + ((0, 0),) * (data.ndim - 3)
                data = np.pad(data, padding, mode='constant', constant_values=(0, padval))
                
                # tile the filters into an image
                data = data.reshape((n, n) + data.shape[1:]).transpose((0, 2, 1, 3) + tuple(range(4, data.ndim + 1)))
                data = data.reshape((n * data.shape[1], n * data.shape[3]) + data.shape[4:])

                data -= data.min()
                data /= data.max()

                cv2.imwrite(curweight_imgpath, data * 255)


            elif len(weightdata.shape)==2:
                imgfname = 'w%02d_fc.jpg' % (idx)
                curweight_imgpath = os.path.join(imgpath, imgfname)
                curweight_relimgpath = os.path.join('imgs', imgfname)
                f.write('<img src="%s" />' % curweight_relimgpath)
                data = weightdata
                data -= data.min()
                data /= data.max()
                cv2.imwrite(curweight_imgpath, data * 255)
            f.write('<hr/>')

    f.write('</body>') 
    f.write('</html>') 
    f.close()

def do_spec(args):
    global CAFFE_ROOT
    fname = os.path.join(CAFFE_ROOT, 'src/caffe/proto/caffe.proto')
    os.system('vim %s' % fname)
    return

def do_latest(args):
    if len(args) < 1:
        print ("Usage: [parameter dir]")
        return

    root = args[0]
    maxnum = -1
    for fname in os.listdir(root):
        if not fname.endswith('.caffemodel'): continue
        curnum = int(fname.split('.')[0].split('_')[-1])
        if curnum > maxnum: maxnum = curnum

    print ("remove all files other than the latest weight? (type \"yes\")")
    if raw_input().strip().lower() != 'yes':
        return

    for fname in os.listdir(root):
        if not fname.split('.')[-1] in ['caffemodel', 'solverstate']: continue
        curnum = int(fname.split('.')[0].split('_')[-1])
        if curnum != maxnum: 
            print ("Removing %s" % fname)
            os.remove(os.path.join(root, fname))

        print ("Done")

def do_countdb(args):
    import lmdb
    if len(args) < 1:
        print ("Usage: [lmdb path]")
        return

    env = lmdb.open(args[0], readonly=True)
    print (env.stat()['entries'])
    env.close()


COMMANDS=[
    ('clean',do_clean),
    ('dataset' ,do_dataset),
    ('ipython' ,do_ipython),
    ('mean'    ,do_mean),
    ('train'   ,do_train),
    ('resume'  ,do_resume),
    ('pltblob' ,do_pltblob),
    ('pltloss' ,do_pltloss),
    ('pltmulti' ,do_pltmulti),
    ('pltacc' ,do_pltacc),
    ('loadstate' ,do_loadstate),
    ('draw', do_draw),
    ('eval', do_eval),
    ('extract', do_extract),
    ('pack', do_pack),
    ('visweight', do_visweight),
    ('visevol', do_visevol),
    ('evolmovie', do_evolmovie),
    ('spec', do_spec),
    ('bench', do_bench),
    ('latest', do_latest),
    ('countdb', do_countdb),
]


def main():
    if len(sys.argv)<2:
        print ("Usage: %s [command]" % sys.argv[0])
        for cmd, fun in COMMANDS:
            print ("   %s" % cmd)
        exit(0)

    for cmd, fun in COMMANDS:
        if sys.argv[1].strip().lower() == cmd:
            fun(sys.argv[2:])
            exit(0)

    print ("Invalid command, availible commands are:")
    for cmd, fun in COMMANDS:
        print ("   %s" % cmd)

if __name__=="__main__":
    main()

